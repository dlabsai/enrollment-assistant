from io import BytesIO
from uuid import uuid4

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.rbac import (
    PermissionKey,
    SystemGroupSlug,
    get_group_for_slug,
    replace_user_permission_overrides,
)
from app.core.security import get_password_hash
from app.main import app
from app.models import Conversation, Message, User
from app.models import Rating as MessageRating
from tests.api.auth_helpers import authenticate_client


async def _create_user(
    session: AsyncSession, *, group_slug: SystemGroupSlug, email_prefix: str
) -> User:
    group = await get_group_for_slug(session, group_slug)
    user = User(
        email=f"{email_prefix}-{uuid4()}@example.com",
        name=f"{group_slug.value.title()} User",
        password_hash=get_password_hash("StrongPassword123"),
        is_active=True,
        group_id=group.id,
    )
    session.add(user)
    await session.commit()
    await session.refresh(user)
    return user


@pytest.mark.asyncio
async def test_message_feedback_routes_support_create_update_list_and_delete(
    transactional_session: AsyncSession,
) -> None:
    owner = await _create_user(
        transactional_session, group_slug=SystemGroupSlug.USER, email_prefix="feedback-owner"
    )
    admin = await _create_user(
        transactional_session, group_slug=SystemGroupSlug.ADMIN, email_prefix="feedback-admin"
    )

    conversation = Conversation(
        title="Feedback test", user=False, project="demo", user_id=owner.id, is_public=False
    )
    transactional_session.add(conversation)
    await transactional_session.flush()

    user_message = Message(role="user", content="Need help", conversation=conversation)
    transactional_session.add(user_message)
    await transactional_session.flush()

    assistant_message = Message(
        role="assistant",
        content="Here is some help",
        conversation=conversation,
        parent_id=user_message.id,
    )
    transactional_session.add(assistant_message)
    await transactional_session.flush()
    user_message.active_child = assistant_message
    await transactional_session.commit()

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as owner_client:
        authenticate_client(owner_client, owner.id)
        create_response = await owner_client.post(
            f"/api/conversations/messages/{assistant_message.id}/feedback",
            json={"rating": MessageRating.THUMBS_UP.value, "text": "Helpful"},
        )
        update_response = await owner_client.post(
            f"/api/conversations/messages/{assistant_message.id}/feedback",
            json={"rating": MessageRating.THUMBS_DOWN.value, "text": "Needs work"},
        )
        owner_list_response = await owner_client.get(
            f"/api/conversations/messages/{assistant_message.id}/feedback"
        )
        owner_detail_response = await owner_client.get(f"/api/conversations/{conversation.id}")

    assert create_response.status_code == 200
    created_feedback = create_response.json()
    assert created_feedback["rating"] == MessageRating.THUMBS_UP.value
    assert created_feedback["text"] == "Helpful"
    assert created_feedback["is_current_user"] is True

    assert update_response.status_code == 200
    updated_feedback = update_response.json()
    assert updated_feedback["id"] == created_feedback["id"]
    assert updated_feedback["rating"] == MessageRating.THUMBS_DOWN.value
    assert updated_feedback["text"] == "Needs work"
    assert updated_feedback["is_current_user"] is True

    assert owner_list_response.status_code == 200
    owner_feedback_items = owner_list_response.json()
    assert len(owner_feedback_items) == 1
    assert owner_feedback_items[0]["is_current_user"] is True

    assert owner_detail_response.status_code == 200
    owner_detail_feedback = owner_detail_response.json()["messages"][1]["feedback"]
    assert owner_detail_feedback[0]["id"] == created_feedback["id"]
    assert owner_detail_feedback[0]["is_current_user"] is True

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as admin_client:
        authenticate_client(admin_client, admin.id)
        admin_list_response = await admin_client.get(
            f"/api/conversations/messages/{assistant_message.id}/feedback"
        )
        admin_tree_response = await admin_client.get(f"/api/conversations/{conversation.id}/tree")
        admin_export_response = await admin_client.get(
            "/api/feedback/export",
            params={
                "browser_time_zone": "UTC",
                "browser_locale": "en-US",
                "message_url_base": "https://internal.example/app",
            },
        )

    assert admin_list_response.status_code == 403
    assert admin_tree_response.status_code == 403
    assert admin_export_response.status_code == 200
    assert admin_export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "feedback-" in admin_export_response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(admin_export_response.content), read_only=True)
    worksheet = workbook.active
    assert worksheet is not None
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "Thumbs",
        "Feedback text",
        "User message",
        "Assistant message",
        "Transcript URL",
        "Chat",
        "Chat user name",
        "Chat user email",
        "Feedback by name",
        "Feedback by email",
        "Created",
    )
    assert rows[1][0] == "Down"
    assert rows[1][1] == "Needs work"
    assert rows[1][2] == "Need help"
    assert rows[1][3] == "Here is some help"
    assert rows[1][4] == (
        f"https://internal.example/app#/chats/{conversation.id}?message={assistant_message.id}"
    )
    assert rows[1][5] == "Feedback test"
    assert rows[1][6] == owner.name
    assert rows[1][7] == owner.email
    assert rows[1][8] == owner.name
    assert rows[1][9] == owner.email
    assert isinstance(rows[1][10], str)
    assert ":" in rows[1][10]
    workbook.close()

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as owner_client:
        authenticate_client(owner_client, owner.id)
        delete_response = await owner_client.delete(
            f"/api/conversations/messages/feedback/{created_feedback['id']}"
        )
        feedback_after_delete_response = await owner_client.get(
            f"/api/conversations/messages/{assistant_message.id}/feedback"
        )

    assert delete_response.status_code == 204
    assert feedback_after_delete_response.status_code == 200
    assert feedback_after_delete_response.json() == []


@pytest.mark.asyncio
async def test_feedback_list_filters_support_owner_group_shortcuts(
    transactional_session: AsyncSession,
) -> None:
    viewer = await _create_user(
        transactional_session, group_slug=SystemGroupSlug.DEV, email_prefix="feedback-viewer"
    )
    peer_user = await _create_user(
        transactional_session, group_slug=SystemGroupSlug.USER, email_prefix="feedback-user-owner"
    )
    peer_admin = await _create_user(
        transactional_session, group_slug=SystemGroupSlug.ADMIN, email_prefix="feedback-admin-owner"
    )
    peer_dev = await _create_user(
        transactional_session, group_slug=SystemGroupSlug.DEV, email_prefix="feedback-dev-owner"
    )

    user_conversation = Conversation(
        title="User chat", user=False, project="demo", user_id=peer_user.id, is_public=False
    )
    admin_conversation = Conversation(
        title="Admin chat", user=False, project="demo", user_id=peer_admin.id, is_public=False
    )
    dev_conversation = Conversation(
        title="Dev chat", user=False, project="demo", user_id=peer_dev.id, is_public=False
    )
    transactional_session.add_all([user_conversation, admin_conversation, dev_conversation])
    await transactional_session.flush()

    user_message = Message(role="user", content="User prompt", conversation=user_conversation)
    admin_message = Message(role="user", content="Admin prompt", conversation=admin_conversation)
    dev_message = Message(role="user", content="Dev prompt", conversation=dev_conversation)
    transactional_session.add_all([user_message, admin_message, dev_message])
    await transactional_session.flush()

    user_assistant = Message(
        role="assistant",
        content="User answer",
        parent_id=user_message.id,
        conversation=user_conversation,
    )
    admin_assistant = Message(
        role="assistant",
        content="Admin answer",
        parent_id=admin_message.id,
        conversation=admin_conversation,
    )
    dev_assistant = Message(
        role="assistant",
        content="Dev answer",
        parent_id=dev_message.id,
        conversation=dev_conversation,
    )
    transactional_session.add_all([user_assistant, admin_assistant, dev_assistant])
    await transactional_session.commit()

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as user_client:
        authenticate_client(user_client, peer_user.id)
        user_feedback_response = await user_client.post(
            f"/api/conversations/messages/{user_assistant.id}/feedback",
            json={"rating": MessageRating.THUMBS_UP.value, "text": "From user"},
        )
        assert user_feedback_response.status_code == 200

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as admin_client:
        authenticate_client(admin_client, peer_admin.id)
        admin_feedback_response = await admin_client.post(
            f"/api/conversations/messages/{admin_assistant.id}/feedback",
            json={"rating": MessageRating.THUMBS_UP.value, "text": "From admin"},
        )
        assert admin_feedback_response.status_code == 200

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as dev_feedback_client:
        authenticate_client(dev_feedback_client, peer_dev.id)
        dev_feedback_response = await dev_feedback_client.post(
            f"/api/conversations/messages/{dev_assistant.id}/feedback",
            json={"rating": MessageRating.THUMBS_UP.value, "text": "From dev"},
        )
        assert dev_feedback_response.status_code == 200

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as viewer_client:
        authenticate_client(viewer_client, viewer.id)
        staff_feedback_response = await viewer_client.get(
            "/api/feedback", params={"limit": 20, "offset": 0, "user_group": "staff"}
        )
        devs_feedback_response = await viewer_client.get(
            "/api/feedback", params={"limit": 20, "offset": 0, "user_group": "devs"}
        )
        conflict_response = await viewer_client.get(
            "/api/feedback",
            params={
                "limit": 20,
                "offset": 0,
                "user_group": "staff",
                "user_email": peer_admin.email,
            },
        )
        staff_export_response = await viewer_client.get(
            "/api/feedback/export",
            params={
                "browser_time_zone": "UTC",
                "browser_locale": "en-US",
                "message_url_base": "https://internal.example/app",
                "user_group": "staff",
            },
        )
        devs_export_response = await viewer_client.get(
            "/api/feedback/export",
            params={
                "browser_time_zone": "UTC",
                "browser_locale": "en-US",
                "message_url_base": "https://internal.example/app",
                "user_group": "devs",
            },
        )
        conflict_export_response = await viewer_client.get(
            "/api/feedback/export",
            params={
                "browser_time_zone": "UTC",
                "browser_locale": "en-US",
                "message_url_base": "https://internal.example/app",
                "user_group": "staff",
                "user_email": peer_admin.email,
            },
        )

    assert staff_feedback_response.status_code == 200
    staff_conversation_ids = {
        item["conversation_id"] for item in staff_feedback_response.json()["items"]
    }
    assert staff_conversation_ids == {str(user_conversation.id), str(admin_conversation.id)}

    assert devs_feedback_response.status_code == 200
    dev_conversation_ids = {
        item["conversation_id"] for item in devs_feedback_response.json()["items"]
    }
    assert dev_conversation_ids == {str(dev_conversation.id)}

    assert conflict_response.status_code == 400
    assert conflict_response.json().get("detail") == "Specify only one of user_email or user_group"

    assert staff_export_response.status_code == 200
    assert staff_export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "feedback-" in staff_export_response.headers["content-disposition"]
    staff_workbook = load_workbook(BytesIO(staff_export_response.content), read_only=True)
    staff_sheet = staff_workbook.active
    assert staff_sheet is not None
    staff_rows = list(staff_sheet.iter_rows(values_only=True))
    staff_export_conversation_ids = {
        str(row[4]).split("/chats/")[1].split("?")[0] for row in staff_rows[1:]
    }
    assert staff_export_conversation_ids == {str(user_conversation.id), str(admin_conversation.id)}
    staff_workbook.close()

    assert devs_export_response.status_code == 200
    assert devs_export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    devs_workbook = load_workbook(BytesIO(devs_export_response.content), read_only=True)
    devs_sheet = devs_workbook.active
    assert devs_sheet is not None
    devs_rows = list(devs_sheet.iter_rows(values_only=True))
    devs_export_conversation_ids = {
        str(row[4]).split("/chats/")[1].split("?")[0] for row in devs_rows[1:]
    }
    assert devs_export_conversation_ids == {str(dev_conversation.id)}
    devs_workbook.close()

    assert conflict_export_response.status_code == 400
    assert (
        conflict_export_response.json().get("detail")
        == "Specify only one of user_email or user_group"
    )


@pytest.mark.asyncio
async def test_feedback_list_and_export_require_visible_owner_group_for_user_group_filter(
    transactional_session: AsyncSession,
) -> None:
    viewer = await _create_user(
        transactional_session,
        group_slug=SystemGroupSlug.ADMIN,
        email_prefix="feedback-no-scope-viewer",
    )
    peer_user = await _create_user(
        transactional_session,
        group_slug=SystemGroupSlug.USER,
        email_prefix="feedback-no-scope-user",
    )
    peer_admin = await _create_user(
        transactional_session,
        group_slug=SystemGroupSlug.ADMIN,
        email_prefix="feedback-no-scope-admin",
    )
    peer_dev = await _create_user(
        transactional_session, group_slug=SystemGroupSlug.DEV, email_prefix="feedback-no-scope-dev"
    )

    user_conversation = Conversation(
        title="No-scope user chat",
        user=False,
        project="demo",
        user_id=peer_user.id,
        is_public=False,
    )
    admin_conversation = Conversation(
        title="No-scope admin chat",
        user=False,
        project="demo",
        user_id=peer_admin.id,
        is_public=False,
    )
    dev_conversation = Conversation(
        title="No-scope dev chat", user=False, project="demo", user_id=peer_dev.id, is_public=False
    )
    transactional_session.add_all([user_conversation, admin_conversation, dev_conversation])
    await transactional_session.flush()

    user_message = Message(role="user", content="User prompt", conversation=user_conversation)
    admin_message = Message(role="user", content="Admin prompt", conversation=admin_conversation)
    dev_message = Message(role="user", content="Dev prompt", conversation=dev_conversation)
    transactional_session.add_all([user_message, admin_message, dev_message])
    await transactional_session.flush()

    user_assistant = Message(
        role="assistant",
        content="User answer",
        parent_id=user_message.id,
        conversation=user_conversation,
    )
    admin_assistant = Message(
        role="assistant",
        content="Admin answer",
        parent_id=admin_message.id,
        conversation=admin_conversation,
    )
    dev_assistant = Message(
        role="assistant",
        content="Dev answer",
        parent_id=dev_message.id,
        conversation=dev_conversation,
    )
    transactional_session.add_all([user_assistant, admin_assistant, dev_assistant])
    await transactional_session.commit()

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as user_client:
        authenticate_client(user_client, peer_user.id)
        user_feedback_response = await user_client.post(
            f"/api/conversations/messages/{user_assistant.id}/feedback",
            json={"rating": MessageRating.THUMBS_UP.value, "text": "From user"},
        )
        assert user_feedback_response.status_code == 200

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as admin_client:
        authenticate_client(admin_client, peer_admin.id)
        admin_feedback_response = await admin_client.post(
            f"/api/conversations/messages/{admin_assistant.id}/feedback",
            json={"rating": MessageRating.THUMBS_UP.value, "text": "From admin"},
        )
        assert admin_feedback_response.status_code == 200

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as dev_feedback_client:
        authenticate_client(dev_feedback_client, peer_dev.id)
        dev_feedback_response = await dev_feedback_client.post(
            f"/api/conversations/messages/{dev_assistant.id}/feedback",
            json={"rating": MessageRating.THUMBS_UP.value, "text": "From dev"},
        )
        assert dev_feedback_response.status_code == 200

    await replace_user_permission_overrides(
        transactional_session,
        viewer,
        {
            PermissionKey.ACCESS_CHATS: True,
            PermissionKey.CHATS_VIEW_OWN: False,
            PermissionKey.CHATS_VIEW_USERS: False,
            PermissionKey.CHATS_VIEW_ADMINS: False,
            PermissionKey.CHATS_VIEW_DEVS: False,
        },
    )

    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://testserver"
    ) as viewer_client:
        authenticate_client(viewer_client, viewer.id)
        staff_feedback = await viewer_client.get(
            "/api/feedback", params={"limit": 20, "offset": 0, "user_group": "staff"}
        )
        devs_feedback = await viewer_client.get(
            "/api/feedback", params={"limit": 20, "offset": 0, "user_group": "devs"}
        )
        staff_export = await viewer_client.get(
            "/api/feedback/export",
            params={
                "browser_time_zone": "UTC",
                "browser_locale": "en-US",
                "message_url_base": "https://internal.example/app",
                "user_group": "staff",
            },
        )
        devs_export = await viewer_client.get(
            "/api/feedback/export",
            params={
                "browser_time_zone": "UTC",
                "browser_locale": "en-US",
                "message_url_base": "https://internal.example/app",
                "user_group": "devs",
            },
        )

    assert staff_feedback.status_code == 200
    assert staff_feedback.json()["items"] == []
    assert staff_feedback.json()["total"] == 0

    assert devs_feedback.status_code == 200
    assert devs_feedback.json()["items"] == []
    assert devs_feedback.json()["total"] == 0

    assert staff_export.status_code == 200
    assert staff_export.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    staff_workbook = load_workbook(BytesIO(staff_export.content), read_only=True)
    staff_sheet = staff_workbook.active
    assert staff_sheet is not None
    staff_rows = list(staff_sheet.iter_rows(values_only=True))
    assert len(staff_rows) == 1
    staff_workbook.close()

    assert devs_export.status_code == 200
    assert devs_export.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    devs_workbook = load_workbook(BytesIO(devs_export.content), read_only=True)
    devs_sheet = devs_workbook.active
    assert devs_sheet is not None
    devs_rows = list(devs_sheet.iter_rows(values_only=True))
    assert len(devs_rows) == 1
    devs_workbook.close()
