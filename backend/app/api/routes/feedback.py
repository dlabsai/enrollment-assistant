from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime, tzinfo
from io import BytesIO
from typing import Annotated, Any, Literal, cast
from urllib.parse import quote
from uuid import UUID  # noqa: TC003
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from babel.core import Locale, UnknownLocaleError
from babel.dates import format_datetime
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import false, func, or_, select
from sqlalchemy.orm import aliased

from app.api.deps import CurrentUser, SessionDep
from app.api.routes.owner_group_filter import (
    OwnerGroup,
    build_owner_group_filter,
    validate_exclusive_user_filters,
)
from app.core.rbac import (
    PermissionKey,
    get_allowed_chat_owner_group_slugs,
    get_effective_permission_map,
)
from app.models import Conversation, Message, MessageFeedback, RbacGroup, User
from app.models import Rating as MessageRating

router = APIRouter(prefix="/feedback", tags=["feedback"])


_XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_EXPORT_HEADERS = (
    "Thumbs",
    "Feedback text",
    "User message",
    "Assistant message",
    "Transcript URL",
    "Chat",
    "Chat user name",
    "Chat user email",
    "Feedback by name",
    "Feedback by email",
    "Created",
)
_EXPORT_COLUMN_WIDTHS = (14, 36, 54, 54, 64, 32, 24, 32, 24, 32, 22)

_PREVIEW_MAX_LENGTH = 160


class FeedbackListItem(BaseModel):
    id: UUID
    message_id: UUID
    conversation_id: UUID
    rating: MessageRating
    text: str | None = None
    message_role: str
    message_preview: str
    conversation_title: str | None = None
    conversation_summary: str | None = None
    is_public: bool
    conversation_user_name: str | None = None
    conversation_user_email: str | None = None
    feedback_user_name: str
    feedback_user_email: str
    created_at: datetime
    updated_at: datetime


class FeedbackListPage(BaseModel):
    items: list[FeedbackListItem]
    total: int


@dataclass(frozen=True)
class FeedbackQueryItem:
    id: UUID
    message_id: UUID
    conversation_id: UUID
    rating: MessageRating
    text: str | None
    message_role: str
    message_content: str
    user_message_content: str | None
    conversation_title: str | None
    conversation_summary: str | None
    is_public: bool
    conversation_user_name: str | None
    conversation_user_email: str | None
    feedback_user_name: str
    feedback_user_email: str
    created_at: datetime
    updated_at: datetime


def _is_admin_user(current_user: CurrentUser) -> bool:
    return current_user.group.slug in {"admin", "dev"}


def _format_preview(content: str) -> str:
    normalized = " ".join(content.split())
    if len(normalized) > _PREVIEW_MAX_LENGTH:
        return normalized[:_PREVIEW_MAX_LENGTH] + "..."
    return normalized


def _internal_visibility_condition(
    current_user: CurrentUser, *, permission_map: dict[PermissionKey, bool]
) -> Any:
    conditions: list[Any] = []

    if permission_map.get(PermissionKey.CHATS_VIEW_OWN, False):
        conditions.append(Conversation.user_id == current_user.id)

    allowed_group_slugs = get_allowed_chat_owner_group_slugs(permission_map)
    if allowed_group_slugs:
        conditions.append(RbacGroup.slug.in_(sorted(allowed_group_slugs)))

    if not conditions:
        return false()

    return or_(*conditions)


def _get_platform_scope(current_user: CurrentUser, platform: str | None) -> tuple[bool, bool]:
    if platform is not None and platform not in {"internal", "public"}:
        raise HTTPException(status_code=400, detail="Invalid platform")

    can_view_public = _is_admin_user(current_user)
    if platform == "public" and not can_view_public:
        raise HTTPException(status_code=403, detail="Access denied")

    include_internal = platform in (None, "internal")
    include_public = can_view_public and platform in (None, "public")
    return include_internal, include_public


async def _get_feedback_permission_map(
    session: SessionDep, current_user: CurrentUser
) -> dict[PermissionKey, bool]:
    permission_map = await get_effective_permission_map(session, current_user)
    if not permission_map.get(PermissionKey.ACCESS_CHATS, False):
        raise HTTPException(status_code=403, detail="Access denied")
    return permission_map


def _build_feedback_base_stmt(
    current_user: CurrentUser,
    *,
    permission_map: dict[PermissionKey, bool],
    platform: Literal["internal", "public"] | None,
    rating: MessageRating | None,
    search: str | None,
    user_email: str | None,
    user_group: OwnerGroup | None,
    start: datetime | None,
    end: datetime | None,
) -> Any:
    include_internal, include_public = _get_platform_scope(current_user, platform)
    internal_visibility_condition = _internal_visibility_condition(
        current_user, permission_map=permission_map
    )

    # SQLAlchemy aliases are needed because feedback author and conversation owner are both users.
    feedback_user_alias = aliased(User)
    owner_user_alias = aliased(User)
    user_message_alias = aliased(Message)

    conversation_user_name = owner_user_alias.name.label("conversation_user_name")
    conversation_user_email = owner_user_alias.email.label("conversation_user_email")

    base_stmt = (
        select(
            MessageFeedback,
            Message,
            Conversation,
            feedback_user_alias.name.label("feedback_user_name"),
            feedback_user_alias.email.label("feedback_user_email"),
            conversation_user_name,
            conversation_user_email,
            user_message_alias.content.label("user_message_content"),
        )
        .join(Message, MessageFeedback.message_id == Message.id)
        .join(Conversation, Message.conversation_id == Conversation.id)
        .join(feedback_user_alias, MessageFeedback.user_id == feedback_user_alias.id)
        .outerjoin(
            user_message_alias,
            (Message.parent_id == user_message_alias.id) & (user_message_alias.role == "user"),
        )
        .outerjoin(owner_user_alias, Conversation.user_id == owner_user_alias.id)
        .outerjoin(RbacGroup, owner_user_alias.group_id == RbacGroup.id)
    )

    platform_conditions: list[Any] = []
    if include_internal:
        platform_conditions.append(
            Conversation.is_public.is_(False) & internal_visibility_condition
        )
    if include_public:
        platform_conditions.append(Conversation.is_public.is_(True))
    if platform_conditions:
        base_stmt = base_stmt.where(or_(*platform_conditions))
    base_stmt = base_stmt.where(Conversation.kind == "chat")

    if rating is not None:
        base_stmt = base_stmt.where(MessageFeedback.rating == rating)
    if start is not None:
        base_stmt = base_stmt.where(MessageFeedback.created_at >= start)
    if end is not None:
        base_stmt = base_stmt.where(MessageFeedback.created_at <= end)
    if search is not None and search.strip() != "":
        pattern = f"%{search.strip()}%"
        base_stmt = base_stmt.where(
            or_(
                MessageFeedback.text.ilike(pattern),
                Message.content.ilike(pattern),
                Conversation.title.ilike(pattern),
                Conversation.summary.ilike(pattern),
                feedback_user_alias.name.ilike(pattern),
                feedback_user_alias.email.ilike(pattern),
                owner_user_alias.name.ilike(pattern),
                owner_user_alias.email.ilike(pattern),
                user_message_alias.content.ilike(pattern),
            )
        )
    validate_exclusive_user_filters(user_email=user_email, user_group=user_group)

    if user_email is not None and user_email.strip() != "":
        normalized_email = user_email.strip()
        user_conditions: list[Any] = []
        if include_internal:
            user_conditions.append(owner_user_alias.email == normalized_email)
        base_stmt = base_stmt.where(or_(*user_conditions) if user_conditions else false())

    return build_owner_group_filter(
        base_stmt,
        owner_group=user_group,
        include_internal=include_internal,
        permission_map=permission_map,
    )


def _sort_feedback_stmt(stmt: Any, *, sort_by: str, descending: bool) -> Any:
    sort_map: dict[str, Any] = {
        "created_at": MessageFeedback.created_at,
        "updated_at": MessageFeedback.updated_at,
        "rating": MessageFeedback.rating,
        "conversation_title": Conversation.title,
    }
    sort_column = sort_map.get(sort_by, MessageFeedback.created_at)
    return stmt.order_by(sort_column.desc() if descending else sort_column.asc())


def _row_to_feedback_query_item(row: Any) -> FeedbackQueryItem:
    (
        feedback,
        message,
        conversation,
        feedback_user_name,
        feedback_user_email,
        conversation_user_name_value,
        conversation_user_email_value,
        user_message_content,
    ) = row

    return FeedbackQueryItem(
        id=feedback.id,
        message_id=message.id,
        conversation_id=conversation.id,
        rating=feedback.rating,
        text=feedback.text,
        message_role=message.role,
        user_message_content=user_message_content,
        message_content=message.content,
        conversation_title=conversation.title,
        conversation_summary=conversation.summary,
        is_public=conversation.is_public,
        conversation_user_name=conversation_user_name_value,
        conversation_user_email=conversation_user_email_value,
        feedback_user_name=feedback_user_name,
        feedback_user_email=feedback_user_email,
        created_at=feedback.created_at,
        updated_at=feedback.updated_at,
    )


def _query_item_to_feedback_list_item(item: FeedbackQueryItem) -> FeedbackListItem:
    return FeedbackListItem(
        id=item.id,
        message_id=item.message_id,
        conversation_id=item.conversation_id,
        rating=item.rating,
        text=item.text,
        message_role=item.message_role,
        message_preview=_format_preview(item.message_content),
        conversation_title=item.conversation_title,
        conversation_summary=item.conversation_summary,
        is_public=item.is_public,
        conversation_user_name=item.conversation_user_name,
        conversation_user_email=item.conversation_user_email,
        feedback_user_name=item.feedback_user_name,
        feedback_user_email=item.feedback_user_email,
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


def _safe_excel_text(value: str | None) -> str:
    if value is None:
        return ""
    stripped = value.lstrip()
    if stripped.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _feedback_rating_label(rating: MessageRating) -> str:
    return "Down" if rating == MessageRating.THUMBS_DOWN else "Up"


def _resolve_browser_timezone(time_zone: str) -> tzinfo:
    try:
        return ZoneInfo(time_zone)
    except ZoneInfoNotFoundError:
        return UTC


def _normalize_browser_datetime(value: datetime, *, time_zone: tzinfo) -> datetime:

    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    return value.astimezone(time_zone)


def _normalize_browser_locale(value: str | None) -> str:
    if value is None:
        return "en_US"
    normalized = value.replace("-", "_").strip()
    if normalized == "":
        return "en_US"
    try:
        return str(Locale.parse(normalized))
    except UnknownLocaleError:
        return "en_US"
    except ValueError:
        return "en_US"


def _build_browser_datetime_format(locale: str | None) -> tuple[str, str]:
    locale_name = _normalize_browser_locale(locale)
    locale_obj = Locale.parse(locale_name)
    format_pattern = "MMM d, h:mm a"
    if "a" not in locale_obj.time_formats["short"].pattern:
        format_pattern = "MMM d, HH:mm"
    return locale_name, format_pattern


def _format_browser_table_timestamp(
    value: datetime, *, time_zone: tzinfo, locale_name: str, format_pattern: str
) -> str:
    localized = _normalize_browser_datetime(value, time_zone=time_zone)
    return format_datetime(localized, format_pattern, locale=locale_name)


def _build_message_url(message_url_base: str, conversation_id: UUID, message_id: UUID) -> str:
    base = message_url_base.split("#", maxsplit=1)[0]
    conversation_path = quote(str(conversation_id), safe="")
    message_query = quote(str(message_id), safe="")
    return f"{base}#/chats/{conversation_path}?message={message_query}"


def _build_feedback_workbook(
    items: list[FeedbackQueryItem],
    *,
    message_url_base: str,
    browser_time_zone: str,
    browser_locale: str,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    if worksheet is None:
        raise RuntimeError("Failed to create feedback export worksheet")
    worksheet.title = "Feedback"
    worksheet.freeze_panes = "A2"

    worksheet.append(_EXPORT_HEADERS)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in cast(tuple[Cell, ...], worksheet[1]):
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    timezone_obj = _resolve_browser_timezone(browser_time_zone)
    locale_name, format_pattern = _build_browser_datetime_format(browser_locale)

    for item in items:
        message_url = _build_message_url(message_url_base, item.conversation_id, item.message_id)
        worksheet.append(
            (
                _feedback_rating_label(item.rating),
                _safe_excel_text(item.text),
                _safe_excel_text(item.user_message_content),
                _safe_excel_text(item.message_content),
                _safe_excel_text(message_url),
                _safe_excel_text(item.conversation_title or "Untitled chat"),
                _safe_excel_text(item.conversation_user_name),
                _safe_excel_text(item.conversation_user_email),
                _safe_excel_text(item.feedback_user_name),
                _safe_excel_text(item.feedback_user_email),
                _format_browser_table_timestamp(
                    item.created_at,
                    time_zone=timezone_obj,
                    locale_name=locale_name,
                    format_pattern=format_pattern,
                ),
            )
        )
        row_index = worksheet.max_row
        message_url_cell = cast(Cell, worksheet.cell(row=row_index, column=5))
        message_url_cell.hyperlink = message_url
        message_url_cell.style = "Hyperlink"
        for cell in cast(tuple[Cell, ...], worksheet[row_index]):
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for column_index, width in enumerate(_EXPORT_COLUMN_WIDTHS, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    worksheet.auto_filter.ref = worksheet.dimensions

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _feedback_export_filename() -> str:
    date = datetime.now(UTC).date().isoformat()
    return f"feedback-{date}.xlsx"


@router.get("/export")
async def export_feedback(
    session: SessionDep,
    current_user: CurrentUser,
    platform: Annotated[Literal["internal", "public"] | None, Query()] = None,
    rating: Annotated[MessageRating | None, Query()] = None,
    search: Annotated[str | None, Query()] = None,
    user_email: Annotated[str | None, Query()] = None,
    user_group: Annotated[OwnerGroup | None, Query()] = None,
    start: Annotated[datetime | None, Query()] = None,
    end: Annotated[datetime | None, Query()] = None,
    sort_by: Annotated[str, Query()] = "created_at",
    descending: Annotated[bool, Query()] = True,
    message_url_base: Annotated[str, Query()] = "",
    browser_time_zone: Annotated[str, Query()] = "UTC",
    browser_locale: Annotated[str, Query()] = "en-US",
) -> StreamingResponse:
    permission_map = await _get_feedback_permission_map(session, current_user)
    base_stmt = _build_feedback_base_stmt(
        current_user,
        permission_map=permission_map,
        platform=platform,
        rating=rating,
        search=search,
        user_email=user_email,
        user_group=user_group,
        start=start,
        end=end,
    )
    stmt = _sort_feedback_stmt(base_stmt, sort_by=sort_by, descending=descending)
    rows = (await session.execute(stmt)).all()
    items = [_row_to_feedback_query_item(row) for row in rows]
    workbook = _build_feedback_workbook(
        items,
        message_url_base=message_url_base,
        browser_time_zone=browser_time_zone,
        browser_locale=browser_locale,
    )
    filename = _feedback_export_filename()

    return StreamingResponse(
        BytesIO(workbook),
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("", response_model=FeedbackListPage)
async def list_feedback(
    session: SessionDep,
    current_user: CurrentUser,
    platform: Annotated[Literal["internal", "public"] | None, Query()] = None,
    rating: Annotated[MessageRating | None, Query()] = None,
    search: Annotated[str | None, Query()] = None,
    user_email: Annotated[str | None, Query()] = None,
    user_group: Annotated[OwnerGroup | None, Query()] = None,
    start: Annotated[datetime | None, Query()] = None,
    end: Annotated[datetime | None, Query()] = None,
    limit: Annotated[int, Query(ge=1, le=100)] = 20,
    offset: Annotated[int, Query(ge=0)] = 0,
    sort_by: Annotated[str, Query()] = "created_at",
    descending: Annotated[bool, Query()] = True,
) -> FeedbackListPage:
    permission_map = await _get_feedback_permission_map(session, current_user)
    base_stmt = _build_feedback_base_stmt(
        current_user,
        permission_map=permission_map,
        platform=platform,
        rating=rating,
        search=search,
        user_email=user_email,
        user_group=user_group,
        start=start,
        end=end,
    )

    total_stmt = select(func.count()).select_from(base_stmt.subquery())
    total = (await session.execute(total_stmt)).scalar() or 0

    stmt = (
        _sort_feedback_stmt(base_stmt, sort_by=sort_by, descending=descending)
        .offset(offset)
        .limit(limit)
    )

    rows = (await session.execute(stmt)).all()
    items = [_row_to_feedback_query_item(row) for row in rows]
    return FeedbackListPage(
        total=total, items=[_query_item_to_feedback_list_item(item) for item in items]
    )
